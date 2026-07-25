from __future__ import annotations

import argparse
import re
import shutil
import unicodedata
from pathlib import Path
from typing import Dict, Optional

from openpyxl import load_workbook


def normalize(value: object) -> str:
    """Normalize Excel text for reliable exact matching."""
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value)).replace("\xa0", " ")
    text = text.casefold().strip()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def header_map(ws) -> Dict[str, int]:
    return {
        normalize(cell.value): cell.column
        for cell in ws[1]
        if cell.value is not None
    }


def read_tariffs(ws) -> Dict[str, float]:
    headers = header_map(ws)
    name_col = headers.get("intervention name")
    tariff_col = headers.get("negotiated tariff")
    if not name_col or not tariff_col:
        raise ValueError(
            f"Sheet '{ws.title}' must contain 'Intervention Name' and "
            "'Negotiated Tariff' columns."
        )

    tariffs: Dict[str, float] = {}
    for row in range(2, ws.max_row + 1):
        name = normalize(ws.cell(row, name_col).value)
        tariff = ws.cell(row, tariff_col).value
        if name and tariff not in (None, ""):
            tariffs[name] = tariff
    return tariffs


def find_opd_tariff(template_name: object, tariffs: Dict[str, float]) -> Optional[float]:
    key = normalize(template_name)

    # First attempt a true normalized exact match.
    if key in tariffs:
        return tariffs[key]

    # Controlled aliases for the wording used in the supplied sample files.
    if "specialist" in key:
        for source_key, tariff in tariffs.items():
            if "specialist" in source_key:
                return tariff

    if "general" in key and ("practioner" in key or "practitioner" in key):
        preferred_terms = ("fixed fee", "general practitioner", "general practioner")
        for term in preferred_terms:
            for source_key, tariff in tariffs.items():
                if term in source_key:
                    return tariff

    return None


def populate_and_remove_unmatched(output_path: Path, facility_path: Path) -> tuple[int, int]:
    facility_wb = load_workbook(facility_path, data_only=True, read_only=True)
    output_wb = load_workbook(output_path)

    if "Imaging" not in facility_wb.sheetnames or "OPD" not in facility_wb.sheetnames:
        raise ValueError("Facility workbook must contain both 'Imaging' and 'OPD' sheets.")
    if "Imaging" not in output_wb.sheetnames or "OPD" not in output_wb.sheetnames:
        raise ValueError("Template workbook must contain both 'Imaging' and 'OPD' sheets.")

    imaging_tariffs = read_tariffs(facility_wb["Imaging"])
    opd_tariffs = read_tariffs(facility_wb["OPD"])

    # IMAGING: facility Intervention Name matches template Sub-Benefit.
    ws = output_wb["Imaging"]
    headers = header_map(ws)
    sub_benefit_col = headers.get("sub benefit")
    tariff_col = headers.get("negotiated tariff")
    if not sub_benefit_col or not tariff_col:
        raise ValueError("Template Imaging sheet needs 'Sub - Benefit' and 'Negotiated Tariff'.")

    imaging_delete = []
    imaging_kept = 0
    for row in range(2, ws.max_row + 1):
        key = normalize(ws.cell(row, sub_benefit_col).value)
        tariff = imaging_tariffs.get(key)
        if tariff is None:
            imaging_delete.append(row)
        else:
            ws.cell(row, tariff_col).value = tariff
            imaging_kept += 1

    for row in reversed(imaging_delete):
        ws.delete_rows(row, 1)

    # OPD: exact match first, then controlled aliases.
    ws = output_wb["OPD"]
    headers = header_map(ws)
    intervention_col = headers.get("intervention name")
    tariff_col = headers.get("negotiated tariff")
    if not intervention_col or not tariff_col:
        raise ValueError("Template OPD sheet needs 'Intervention Name' and 'Negotiated Tariff'.")

    opd_delete = []
    opd_kept = 0
    for row in range(2, ws.max_row + 1):
        tariff = find_opd_tariff(ws.cell(row, intervention_col).value, opd_tariffs)
        if tariff is None:
            opd_delete.append(row)
        else:
            ws.cell(row, tariff_col).value = tariff
            opd_kept += 1

    for row in reversed(opd_delete):
        ws.delete_rows(row, 1)

    output_wb.save(output_path)
    return imaging_kept, opd_kept


def process_file(facility_file: str | Path, template_file: str | Path, output_dir: str | Path) -> Path:
    facility_path = Path(facility_file)
    template_path = Path(template_file)
    output_dir = Path(output_dir)

    if not facility_path.exists():
        raise FileNotFoundError(f"Facility file not found: {facility_path}")
    if not template_path.exists():
        raise FileNotFoundError(f"Template file not found: {template_path}")

    output_dir.mkdir(parents=True, exist_ok=True)
    # Preserve the uploaded facility filename exactly.
    output_path = output_dir / facility_path.name
    shutil.copy2(template_path, output_path)

    imaging_count, opd_count = populate_and_remove_unmatched(output_path, facility_path)
    print(f"Created: {output_path}")
    print(f"Imaging rows retained: {imaging_count}")
    print(f"OPD rows retained: {opd_count}")
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Populate the Imaging and OPD template from a facility workbook."
    )
    parser.add_argument("facility_file", help="Facility Excel workbook")
    parser.add_argument("template_file", help="Imaging and OPD template workbook")
    parser.add_argument(
        "--output-dir", default="output", help="Folder for completed workbooks"
    )
    args = parser.parse_args()
    process_file(args.facility_file, args.template_file, args.output_dir)


if __name__ == "__main__":
    main()
